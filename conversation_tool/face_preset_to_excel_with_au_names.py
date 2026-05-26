"""表情プリセット形式のテキストファイルを Excel に変換するスクリプト。

入力形式（複数ブロックを連続で記述可能）:
[name]
<200, 1000, 4000>
{AU1, AU2, ..., AU35}

Excel 出力:
- face_presets シート: 1プリセット = 1行で値を一覧表示
- AU対応表 シート: AU番号と部位名の対応を一覧表示

実行例:
    python face_preset_to_excel_with_au_names.py face_presets.txt
    python face_preset_to_excel_with_au_names.py face_presets.txt -o face_presets.xlsx
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

Number = Union[int, float]
AU_COUNT = 35

# ユーザーが確認したロボット表情パラメータの対応表
AU_LABELS: dict[int, str] = {
    1: "左目の上瞼の開き度合い",
    2: "右目の上瞼の開き度合い",
    3: "左目の向く方向",
    4: "右目の向く方向",
    5: "目の位置上下",
    6: "左目の下瞼",
    7: "右目の下瞼",
    8: "左眉左上げ",
    9: "左眉左下げ",
    10: "左眉右側",
    11: "左 眉間",
    12: "右眉右側上げ",
    13: "右眉右側下げ",
    14: "右眉左側上げ",
    15: "右 眉間",
    16: "左口角",
    17: "右口角",
    18: "左頬横",
    19: "左頬下",
    20: "左頬？",
    21: "左頬下の方",
    22: "右頬横",
    23: "右頬下",
    24: "右頬？",
    25: "左頬下の方",
    26: "上唇の尖り",
    27: "下唇の尖り",
    28: "上唇上",
    29: "下唇下",
    30: "鼻上",
    31: "？",
    32: "口の開き度合い",
    33: "首の角度横",
    34: "首の角度縦",
    35: "首の角度回転",
}

BLOCK_PATTERN = re.compile(
    r"\[(?P<name>[^\]\r\n]+)\]\s*"
    r"<(?P<meta>[^>]*)>\s*"
    r"\{(?P<aus>[^}]*)\}",
    flags=re.DOTALL,
)


def parse_number(value: str) -> Number:
    """数値文字列を int または float に変換する。"""
    cleaned = value.strip()
    if not cleaned:
        raise ValueError("空の値が含まれています。")
    try:
        number = float(cleaned)
    except ValueError as exc:
        raise ValueError(f"数値として読めない値があります: {cleaned!r}") from exc
    return int(number) if number.is_integer() else number


def parse_values(text: str) -> list[Number]:
    """カンマ・全角カンマ・改行・空白区切りの値を数値リストにする。"""
    parts = re.split(r"[,，\s]+", text.strip())
    return [parse_number(part) for part in parts if part]


def read_presets(input_path: Path) -> list[list[object]]:
    """入力テキストから [name] / <3値> / {AU35値} のブロックを読み取る。"""
    text = input_path.read_text(encoding="utf-8-sig")
    rows: list[list[object]] = []

    for index, match in enumerate(BLOCK_PATTERN.finditer(text), start=1):
        name = match.group("name").strip()
        meta_values = parse_values(match.group("meta"))
        au_values = parse_values(match.group("aus"))

        if len(meta_values) != 3:
            raise ValueError(
                f"{index}個目のプリセット [{name}] の <> 内の値が "
                f"{len(meta_values)} 個です。3個である必要があります。"
            )
        if len(au_values) != AU_COUNT:
            raise ValueError(
                f"{index}個目のプリセット [{name}] の AU 値が "
                f"{len(au_values)} 個です。{AU_COUNT} 個である必要があります。"
            )

        rows.append([name, *meta_values, *au_values])

    if not rows:
        raise ValueError(
            "読み取れるプリセットがありませんでした。"
            "[name]、<3個の値>、{AU1〜AU35} の形式を確認してください。"
        )

    return rows


def apply_header_style(sheet, header_row: int = 1) -> None:
    """見出し行の共通スタイルを適用する。"""
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel(rows: list[list[object]], output_path: Path) -> None:
    """プリセット一覧とAU対応表を整形済み Excel として保存する。"""
    workbook = Workbook()

    # -------------------------
    # プリセット値一覧シート
    # -------------------------
    preset_sheet = workbook.active
    preset_sheet.title = "face_presets"

    headers = ["name", "param_1", "param_2", "param_3"] + [
        f"AU{i}\n{AU_LABELS[i]}" for i in range(1, AU_COUNT + 1)
    ]
    preset_sheet.append(headers)
    for row in rows:
        preset_sheet.append(row)

    apply_header_style(preset_sheet)

    for cell in preset_sheet["A"][1:]:
        cell.alignment = Alignment(horizontal="left", vertical="center")

    preset_sheet.freeze_panes = "B2"
    preset_sheet.auto_filter.ref = preset_sheet.dimensions
    preset_sheet.row_dimensions[1].height = 42
    preset_sheet.column_dimensions["A"].width = 24
    for column_index in range(2, 5):
        preset_sheet.column_dimensions[get_column_letter(column_index)].width = 12
    for column_index in range(5, len(headers) + 1):
        preset_sheet.column_dimensions[get_column_letter(column_index)].width = 18

    # -------------------------
    # AU番号・部位名の対応表シート
    # -------------------------
    mapping_sheet = workbook.create_sheet("AU対応表")
    mapping_sheet.append(["AU番号", "部位名"])
    for au_number, label in AU_LABELS.items():
        mapping_sheet.append([f"AU{au_number}", label])

    apply_header_style(mapping_sheet)
    for row in mapping_sheet.iter_rows(min_row=2):
        row[1].alignment = Alignment(horizontal="left", vertical="center")
    mapping_sheet.freeze_panes = "A2"
    mapping_sheet.auto_filter.ref = mapping_sheet.dimensions
    mapping_sheet.column_dimensions["A"].width = 14
    mapping_sheet.column_dimensions["B"].width = 28

    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="表情プリセットのテキストファイルを、AU部位名付きの Excel 一覧に変換します。"
    )
    parser.add_argument("input", type=Path, help="入力テキストファイルのパス")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="出力 Excel ファイルのパス（省略時は入力ファイルと同名の .xlsx）",
    )
    args = parser.parse_args()

    input_path: Path = args.input
    output_path: Path = args.output or input_path.with_suffix(".xlsx")

    if not input_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {input_path}")

    rows = read_presets(input_path)
    write_excel(rows, output_path)
    print(f"{len(rows)} 個のプリセットを出力しました: {output_path}")


if __name__ == "__main__":
    main()
